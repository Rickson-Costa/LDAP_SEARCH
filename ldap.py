import ldap3
import logging
from ldap3 import Server, Connection, SUBTREE, ALL_ATTRIBUTES

# Configuração do logger
logger = logging.getLogger(__name__)

from openpyxl import load_workbook

# Carrega o arquivo Excel
arquivo = "teste.xlsx" # RECEBENDO UMA PLANILHA QUE CONTÉM EM SUA PRIMEIRA COLUNA UMA LISTA DE EMAILS!
wb = load_workbook(arquivo)
sheet = wb.active  # Pega a aba ativa



def consultar(email):
    ldap_server = 'IP'                         #settings.LDAP_IP_SERVER
    ldap_domain = 'DOMINIO'                              #settings.LDAP_DOMAIN_NAME
    
    ldap_base_dn = "OU=Usuarios,DC={},DC=local".format(ldap_domain)

    logger.warning("Tentativa de consulta para {}".format(email))

    # Conecta-se ao servidor LDAP
    server = ldap3.Server('ldap://' + ldap_server)
    conn = ldap3.Connection(server, user='{}\\{}'.format(ldap_domain, 'usuarios'), password='senha', authentication=ldap3.SIMPLE) # logue aqui

    # Tenta conectar
    if conn.bind():
        # Realiza uma pesquisa LDAP para obter os atributos do usuário
        conn.search(search_base=ldap_base_dn, search_filter='(mail={})'.format(email), search_scope=SUBTREE, attributes=['name']) # Condicionais de Pesquisa AQUI!
            
        if len(conn.entries) == 1:
            # Recupera o nome completo do usuário da entrada LDAP
            full_name = conn.entries[0]['name'].value # Pegando Nome Completo Através do Email!
            return full_name

# Percorre os e-mails na coluna A (ignora o cabeçalho na linha 1)
for row in range(2, sheet.max_row + 1):  
    email = sheet[f"A{row}"].value  # Pega o e-mail da célula

    if email:  # Se a célula não estiver vazia
        nome_completo = consultar(email)  # Usa sua função para buscar o nome
        sheet[f"B{row}"] = nome_completo  # Escreve o nome na coluna B

# Salva as alterações
wb.save("teste_atualizado.xlsx")
print("Processo concluído! Os nomes foram preenchidos.")